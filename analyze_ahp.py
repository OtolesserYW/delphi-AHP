"""
AHP 专家问卷数据分析脚本
用法：python analyze_ahp.py <导出的csv路径> <输出xlsx路径>

功能：
1. 解析 Streamlit 问卷后台导出的 CSV（matrices_json / cr_json 为嵌套 JSON 字符串）
2. 为每个层级/每组指标生成一个工作表：展示各专家原始判断矩阵、
   按"行几何平均法"（与你参考文献中 AHP 权重计算公式一致：Wi=(∏aij)^(1/n)后归一化）
   算出的专家局部权重，再用 GEOMEAN 对多位专家的局部权重做几何平均聚合
3. 生成"组合权重汇总"工作表：将二级/三级指标的局部权重逐级相乘，
   得到相对于总目标的最终组合权重（全局权重），并按大小排序
4. 全程使用 Excel 公式（GEOMEAN / SUM / 乘积），而非把计算结果写死，
   这样以后再导出新一轮数据，只需替换原始判断矩阵区域即可重新计算
"""
import sys
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Microsoft YaHei"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = PatternFill("solid", fgColor="D9E2F3")
INPUT_FONT = Font(name=FONT_NAME, color="0000FF")  # 蓝色=硬编码输入（专家原始打分）
CALC_FONT = Font(name=FONT_NAME, color="000000")   # 黑色=公式计算结果
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BOLD = Font(name=FONT_NAME, bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GROUP_ORDER = ["L1", "L2_A", "L2_B", "L2_C",
               "L3_A1", "L3_A2", "L3_A3", "L3_A4", "L3_A5", "L3_A6",
               "L3_B1", "L3_B2", "L3_B3", "L3_C1", "L3_C2", "L3_C3"]

# 组内的父级指标名称（用于汇总表拼接层级关系）
PARENT_OF = {
    "L2_A": "A.病原体与相应疾病风险特征", "L2_B": "B.环境暴露风险", "L2_C": "C.个体风险与健康基础",
    "L3_A1": "A1.病原体基础属性", "L3_A2": "A2.病原体变异与进化潜力", "L3_A3": "A3.传播特性与潜力",
    "L3_A4": "A4.环境存活能力", "L3_A5": "A5.疾病临床严重性", "L3_A6": "A6.临床防治有效性",
    "L3_B1": "B1.布局与通风", "L3_B2": "B2.感染者风险", "L3_B3": "B3.物品与环境污染风险",
    "L3_C1": "C1.工作场景暴露风险", "L3_C2": "C2.个体生理易感性", "L3_C3": "C3.个体防护装备的身体适配性",
}
L2_GROUP_OF_L1_ITEM = {"A.病原体与相应疾病风险特征": "L2_A", "B.环境暴露风险": "L2_B", "C.个体风险与健康基础": "L2_C"}


def cell(ws, row, col, value, font=None, fill=None, align=None, border=BORDER, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or Font(name=FONT_NAME)
    if fill:
        c.fill = fill
    c.alignment = align or Alignment(vertical="center", wrap_text=True)
    c.border = border
    if number_format:
        c.number_format = number_format
    return c


def build_group_sheet(wb, group_key, items, expert_names, matrices, sheet_name):
    """为一个指标组生成工作表，返回 {item: 聚合后归一化权重所在单元格坐标} 供汇总表引用"""
    ws = wb.create_sheet(sheet_name)
    n = len(items)
    n_experts = len(expert_names)
    ws.sheet_view.showGridLines = False

    r = 1
    cell(ws, r, 1, f"{sheet_name}  下属指标两两比较矩阵与权重计算", BOLD)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + n)
    r += 2

    weight_col_start = {}  # expert_idx -> 该专家局部权重列首行(该专家权重列的列号)
    expert_block_start_row = r

    for e_idx, ename in enumerate(expert_names):
        cell(ws, r, 1, f"专家：{ename}", BOLD, fill=GROUP_FILL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + n)
        r += 1
        header_row = r
        cell(ws, r, 1, "指标", HEADER_FONT, fill=HEADER_FILL)
        for j, it in enumerate(items):
            cell(ws, r, 2 + j, it, HEADER_FONT, fill=HEADER_FILL,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        cell(ws, r, 2 + n, "行几何平均", HEADER_FONT, fill=HEADER_FILL,
             align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        cell(ws, r, 3 + n, "归一化局部权重", HEADER_FONT, fill=HEADER_FILL,
             align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        r += 1
        matrix_top = r
        matrix = matrices[e_idx][group_key]["matrix"]
        for i in range(n):
            cell(ws, r, 1, items[i], BOLD)
            for j in range(n):
                cell(ws, r, 2 + j, matrix[i][j], INPUT_FONT, number_format="0.000")
            r += 1
        matrix_bottom = r - 1

        geomean_col = 2 + n
        norm_col = 3 + n
        for i in range(matrix_top, matrix_bottom + 1):
            row_range = f"{get_column_letter(2)}{i}:{get_column_letter(1 + n)}{i}"
            cell(ws, i, geomean_col, f"=GEOMEAN({row_range})", CALC_FONT, number_format="0.0000")
        sum_geomean_cell = f"{get_column_letter(geomean_col)}{matrix_top}:{get_column_letter(geomean_col)}{matrix_bottom}"
        for i in range(matrix_top, matrix_bottom + 1):
            g_ref = f"{get_column_letter(geomean_col)}{i}"
            cell(ws, i, norm_col, f"={g_ref}/SUM({sum_geomean_cell})", CALC_FONT, number_format="0.0000")

        weight_col_start[e_idx] = (matrix_top, norm_col)
        r = matrix_bottom + 2

    # ---- 各专家局部权重汇总 + 几何平均聚合 ----
    cell(ws, r, 1, "各专家局部权重汇总与聚合（几何平均法聚合多位专家意见）", BOLD, fill=GROUP_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3 + n_experts)
    r += 1
    header_row2 = r
    cell(ws, r, 1, "指标", HEADER_FONT, fill=HEADER_FILL)
    for e_idx, ename in enumerate(expert_names):
        cell(ws, r, 2 + e_idx, ename, HEADER_FONT, fill=HEADER_FILL,
             align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    agg_col = 2 + n_experts
    norm_agg_col = 3 + n_experts
    cell(ws, r, agg_col, "几何平均", HEADER_FONT, fill=HEADER_FILL,
         align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    cell(ws, r, norm_agg_col, "归一化组内权重", HEADER_FONT, fill=HEADER_FILL,
         align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    r += 1
    summary_top = r
    for i, it in enumerate(items):
        cell(ws, r, 1, it, BOLD)
        for e_idx in range(n_experts):
            top, norm_col = weight_col_start[e_idx]
            src = f"{get_column_letter(norm_col)}{top + i}"
            cell(ws, r, 2 + e_idx, f"='{ws.title}'!{src}", CALC_FONT, number_format="0.0000")
        expert_range = f"{get_column_letter(2)}{r}:{get_column_letter(1 + n_experts)}{r}"
        cell(ws, r, agg_col, f"=GEOMEAN({expert_range})", CALC_FONT, number_format="0.0000")
        r += 1
    summary_bottom = r - 1
    agg_sum_range = f"{get_column_letter(agg_col)}{summary_top}:{get_column_letter(agg_col)}{summary_bottom}"
    for i in range(summary_top, summary_bottom + 1):
        g_ref = f"{get_column_letter(agg_col)}{i}"
        cell(ws, i, norm_agg_col, f"={g_ref}/SUM({agg_sum_range})", CALC_FONT, number_format="0.0000")

    ws.column_dimensions["A"].width = 30
    for j in range(2, 5 + max(n, n_experts)):
        ws.column_dimensions[get_column_letter(j)].width = 14

    item_cell_ref = {items[i]: f"'{ws.title}'!{get_column_letter(norm_agg_col)}{summary_top + i}" for i in range(n)}
    return item_cell_ref


def build_cr_sheet(wb, groups, items_map, expert_names, crs, sheet_name="一致性检验CR"):
    ws = wb.create_sheet(sheet_name)
    cell(ws, 1, 1, "各专家 · 各组判断矩阵一致性比率 CR（提交时已要求 CR<0.1，此表供复核）", BOLD)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(expert_names))
    r = 3
    cell(ws, r, 1, "指标组", HEADER_FONT, fill=HEADER_FILL)
    for e_idx, ename in enumerate(expert_names):
        cell(ws, r, 2 + e_idx, ename, HEADER_FONT, fill=HEADER_FILL)
    r += 1
    for g in groups:
        cell(ws, r, 1, g)
        for e_idx in range(len(expert_names)):
            v = crs[e_idx].get(g)
            f = CALC_FONT
            fill = None
            if isinstance(v, (int, float)) and v >= 0.1:
                fill = PatternFill("solid", fgColor="FFF2CC")
                f = Font(name=FONT_NAME, color="D97706", bold=True)
            cell(ws, r, 2 + e_idx, round(v, 4) if isinstance(v, (int, float)) else v, f, fill=fill, number_format="0.0000")
        r += 1
    ws.column_dimensions["A"].width = 12
    for j in range(2, 2 + len(expert_names)):
        ws.column_dimensions[get_column_letter(j)].width = 14


def build_expert_sheet(wb, df, sheet_name="专家名单"):
    ws = wb.create_sheet(sheet_name, 0)
    cell(ws, 1, 1, "已收集问卷专家名单", BOLD)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    r = 3
    for j, h in enumerate(["序号", "专家姓名", "轮次", "提交时间"]):
        cell(ws, r, 1 + j, h, HEADER_FONT, fill=HEADER_FILL)
    r += 1
    for _, row in df.iterrows():
        cell(ws, r, 1, int(row["id"]))
        cell(ws, r, 2, row["expert_name"])
        cell(ws, r, 3, int(row["round_no"]))
        cell(ws, r, 4, row["submit_time"])
        r += 1
    for col, w in zip("ABCD", [8, 20, 8, 22]):
        ws.column_dimensions[col].width = w


def build_summary_sheet(wb, item_cell_refs, wb_group_items):
    ws = wb.create_sheet("组合权重汇总", 1)
    cell(ws, 1, 1, "各级指标组合权重（相对总目标的全局权重）汇总", BOLD)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    r = 3
    headers = ["一级指标", "二级指标", "三级指标", "组内局部权重", "组合权重（全局）", "备注"]
    for j, h in enumerate(headers):
        cell(ws, r, 1 + j, h, HEADER_FONT, fill=HEADER_FILL)
    r += 1

    l1_items = wb_group_items["L1"]
    rows_written = []
    for l1_item in l1_items:
        l1_local_ref = item_cell_refs["L1"][l1_item]
        l2_group = L2_GROUP_OF_L1_ITEM[l1_item]
        l2_items = wb_group_items[l2_group]
        for l2_item in l2_items:
            l2_local_ref = item_cell_refs[l2_group][l2_item]
            l2_global_formula = f"={l1_local_ref}*{l2_local_ref}"
            l3_group = [k for k, v in PARENT_OF.items() if v == l2_item and k.startswith("L3_")]
            if l3_group:
                l3_group = l3_group[0]
                l3_items = wb_group_items[l3_group]
                first = True
                for l3_item in l3_items:
                    l3_local_ref = item_cell_refs[l3_group][l3_item]
                    cell(ws, r, 1, l1_item if first else "")
                    cell(ws, r, 2, l2_item if first else "")
                    cell(ws, r, 3, l3_item)
                    cell(ws, r, 4, f"={l3_local_ref}", number_format="0.0000")
                    l2_global_cell = f"$G${r}" if False else None
                    # 组合权重 = L1局部 * L2局部 * L3局部
                    cell(ws, r, 5, f"={l1_local_ref}*{l2_local_ref}*{l3_local_ref}", CALC_FONT, number_format="0.0000")
                    first = False
                    rows_written.append(r)
                    r += 1
            else:
                cell(ws, r, 1, l1_item)
                cell(ws, r, 2, l2_item)
                cell(ws, r, 3, "(无三级指标)")
                cell(ws, r, 5, l2_global_formula, CALC_FONT, number_format="0.0000")
                rows_written.append(r)
                r += 1

    total_row = r + 1
    cell(ws, total_row, 3, "合计", BOLD)
    first_data_row = 4
    last_data_row = r - 1
    cell(ws, total_row, 5, f"=SUM(E{first_data_row}:E{last_data_row})", BOLD, number_format="0.0000")

    for col, w in zip("ABCDEF", [26, 26, 30, 14, 16, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    return first_data_row, last_data_row


def build_workbook(df: pd.DataFrame) -> Workbook:
    """
    核心分析入口：接收一份包含 expert_name / matrices_json / cr_json 等列的
    DataFrame（不管数据来自 CSV 文件还是数据库查询结果，只要列名一致即可），
    返回生成好的 openpyxl Workbook 对象，不涉及任何文件读写。

    这样 Streamlit 网站的管理员面板可以直接把数据库查出来的 df 传进来，
    在内存里生成 Excel 供下载，不需要先导出 CSV 再本地运行脚本。
    """
    df = df.reset_index(drop=True)
    expert_names = df["expert_name"].tolist()
    mats = [json.loads(x) for x in df["matrices_json"]]
    crs = [json.loads(x) for x in df["cr_json"]]

    wb_group_items = {g: mats[0][g]["items"] for g in GROUP_ORDER}

    wb = Workbook()
    wb.remove(wb.active)

    item_cell_refs = {}
    for g in GROUP_ORDER:
        items = wb_group_items[g]
        refs = build_group_sheet(wb, g, items, expert_names, mats, sheet_name=g)
        item_cell_refs[g] = refs

    build_expert_sheet(wb, df)
    build_summary_sheet(wb, item_cell_refs, wb_group_items)
    build_cr_sheet(wb, GROUP_ORDER, wb_group_items, expert_names, crs)

    return wb


def main(csv_path, out_path):
    """命令行入口：仍然保留，本地跑 `python analyze_ahp.py xxx.csv xxx.xlsx` 不受影响。"""
    df = pd.read_csv(csv_path)
    wb = build_workbook(df)
    wb.save(out_path)
    print("saved:", out_path)


if __name__ == "__main__":
    csv_path = sys.argv[1] if len(sys.argv) > 1 else "ahp_responses_round3.csv"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "AHP权重分析结果.xlsx"
    main(csv_path, out_path)
